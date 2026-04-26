# data_loader.py
# Loads any AMOS-format input xlsx and returns clean Python dicts.
# All parsing is done here; agents receive clean typed values, never raw cell strings.
# Brand-agnostic: works for any input file matching the AMOS sheet schema.

import re
import openpyxl


# ─── PARSING HELPERS ─────────────────────────────────────────────────────────

def parse_pct(v):
    """'40%' → 0.40, '-10%' → -0.10, 0.521 (already fraction) → 0.521, None → None"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        # Already a decimal fraction if magnitude ≤ 1
        return f if -1.0 <= f <= 1.0 else f / 100.0
    s = str(v).strip()
    if s.endswith('%'):
        try:
            return float(s[:-1]) / 100.0
        except ValueError:
            return None
    try:
        f = float(s)
        return f if -1.0 <= f <= 1.0 else f / 100.0
    except ValueError:
        return None


def parse_currency(v):
    """'₹2,50,000' → 250000.0, '$1,200' → 1200.0, '50 Cr' → 50.0, 443.4 → 443.4, None → None.
    Strips currency symbols (₹, $), commas, whitespace, and a trailing 'Cr' suffix."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    s = re.sub(r'[₹$,\s]', '', s)
    s = re.sub(r'[Cc][Rr]\.?$', '', s).strip()
    try:
        return float(s)
    except ValueError:
        return None


def parse_score(v):
    """'Low' → 1, 'Medium' → 2, 'High' → 3, None → None"""
    if v is None:
        return None
    s = str(v).strip().lower()
    if s.startswith('low'):
        return 1
    if s.startswith('med'):
        return 2
    if s.startswith('high'):
        return 3
    return None


# ─── MAIN ENTRY POINT ────────────────────────────────────────────────────────

def load_all_data(filename: str) -> dict:
    """Load all 8 data sheets from the xlsx. Returns a dict with keys:
    biz, cohorts, channels, brand, trends, historic, competitors, assumptions."""
    wb = openpyxl.load_workbook(filename, data_only=True)
    return {
        "biz":         _load_biz(wb),
        "cohorts":     _load_cohorts(wb),
        "channels":    _load_channels(wb),
        "brand":       _load_brand(wb),
        "trends":      _load_trends(wb),
        "historic":    _load_historic(wb),
        "competitors": _load_competitors(wb),
        "assumptions": _load_assumptions(wb),
    }


# ─── SHEET LOADERS ───────────────────────────────────────────────────────────

def _load_biz(wb) -> dict:
    """Business Context: col A = field, col B = value. Data rows 4–20."""
    ws = wb['Business Context']
    biz = {}
    for row in ws.iter_rows(min_row=4, max_row=20, values_only=True):
        if row[0] is not None:
            biz[str(row[0]).strip()] = row[1]
    return biz


def _load_cohorts(wb) -> list:
    """Cohorts sheet is transposed: col A = field names, cols B–E = 4 cohorts.
    Row 3 = headers with cohort names. Data rows 4–19."""
    ws = wb['Cohorts']

    # Cohort names from row 3, cols B-E (may contain newlines)
    cohort_names = []
    for c in range(2, 6):
        raw = ws.cell(row=3, column=c).value
        name = str(raw).split('\n')[0].strip() if raw else f"Cohort{c - 1}"
        cohort_names.append(name)

    cohorts = [{"cohort_name": n} for n in cohort_names]

    # (pattern_substring, dict_key, parser_fn)
    field_map = [
        ('Cohort size',                'size',                  lambda v: int(v) if v is not None else None),
        ('Average Order Value',        'aov',                   parse_currency),
        ('% of current',               'revenue_pct',           lambda v: float(v) if isinstance(v, (int, float)) else parse_pct(v)),
        ('Primary channels',           'primary_channels',      lambda v: str(v).strip() if v else None),
        ('Top product preference',     'product_preference',    lambda v: str(v).strip() if v else None),
        ('Price sensitivity',          'price_sensitivity',     lambda v: str(v).strip().split()[0].capitalize() if v else None),
        ('One-line insight',           'insight',               lambda v: str(v).strip() if v else None),
        ('Growth potential',           'growth_potential',      lambda v: str(v).strip().split()[0].capitalize() if v else None),
        ('Competitive intensity',      'competitive_intensity', lambda v: str(v).strip().split()[0].capitalize() if v else None),
        ('Key competitors',            'key_competitors',       lambda v: str(v).strip() if v else None),
        ('Historic YoY growth',        'historic_yoy_growth',   parse_pct),
        ('Acquisition difficulty',     'acquisition_difficulty',lambda v: str(v).strip().split()[0].capitalize() if v else None),
        ('Retention rate',             'retention_rate',        parse_pct),
        ('Estimated LTV',              'ltv',                   parse_currency),
        ('Reachable share',            'reachable_share',       parse_pct),
    ]

    for row in ws.iter_rows(min_row=4, max_row=19, values_only=True):
        field = row[0]
        if field is None:
            continue
        field_str = str(field).strip()
        for pattern, key, parser in field_map:
            if pattern.lower() in field_str.lower():
                for ci in range(4):
                    val = row[ci + 1]
                    try:
                        cohorts[ci][key] = parser(val)
                    except (TypeError, ValueError, IndexError):
                        cohorts[ci][key] = val
                break

    return cohorts


def _load_channels(wb) -> list:
    """Channel Assumptions: row 4 = headers, rows 5–16 = 12 channels.
    Col H (index 7) = Cost per customer reached (₹) — new field."""
    ws = wb['Channel Assumptions']
    channels = []
    for row in ws.iter_rows(min_row=5, max_row=16, values_only=True):
        if row[0] is None:
            continue
        # Effectiveness multiplier stored as '1.40x'
        eff_raw = row[4]
        eff = float(str(eff_raw).replace('x', '').strip()) if eff_raw else None
        # Budget share stored as '15%'
        budget_raw = row[6]
        budget_pct_f = parse_pct(budget_raw)
        budget_pct = int(round(budget_pct_f * 100)) if budget_pct_f is not None else None
        # Cost per customer reached — col H (index 7), integer rupees
        cpc_raw = row[7] if len(row) > 7 else None
        cost_per_reach = int(cpc_raw) if cpc_raw is not None else None
        channels.append({
            "channel_name":             str(row[0]).strip(),
            "cpm_cac":                  str(row[1]).strip() if row[1] else None,
            "conversion_lift":          str(row[2]).strip() if row[2] else None,
            "best_fit_cohort":          str(row[3]).strip() if row[3] else None,
            "effectiveness_multiplier": eff,
            "measurability":            str(row[5]).strip() if row[5] else None,
            "budget_share_pct":         budget_pct,
            "cost_per_customer_reached": cost_per_reach,
        })
    return channels


def _load_brand(wb) -> dict:
    """Brand Guidelines: col A = dimension, col B = value. Data rows 5–13."""
    ws = wb['Brand Guidelines']
    brand = {}
    for row in ws.iter_rows(min_row=5, max_row=13, values_only=True):
        if row[0] is not None:
            brand[str(row[0]).strip()] = row[1]
    return brand


def _load_trends(wb) -> list:
    """Industry Trends: row 4 = headers, rows 5–12 = 8 trends."""
    ws = wb['Industry Trends']
    trends = []
    for row in ws.iter_rows(min_row=5, max_row=12, values_only=True):
        if row[0] is None:
            continue
        trends.append({
            "trend_name":  str(row[0]).strip(),
            "direction":   str(row[1]).strip() if row[1] else None,
            "relevance":   str(row[2]).strip() if row[2] else None,
            "description": str(row[3]).strip() if row[3] else None,
        })
    return trends


def _load_historic(wb) -> list:
    """Historic Sales Data: row 4 = headers, rows 5–8 = 4 fiscal years."""
    ws = wb['Historic Sales Data']
    historic = []
    for row in ws.iter_rows(min_row=5, max_row=8, values_only=True):
        if row[0] is None:
            continue
        historic.append({
            "year":               str(row[0]).strip(),
            "annual_revenue_cr":  parse_currency(row[1]),
            "festive_revenue_cr": parse_currency(row[2]),
            "yoy_growth_pct":     parse_pct(row[3]),
            "best_cohort":        str(row[4]).strip() if row[4] else None,
            "notes":              str(row[5]).strip() if row[5] else None,
        })
    return historic


def _load_competitors(wb) -> list:
    """Competitive Landscape: row 4 = headers, rows 5–14 = 10 competitors."""
    ws = wb['Competitive Landscape']
    competitors = []
    for row in ws.iter_rows(min_row=5, max_row=14, values_only=True):
        if row[0] is None:
            continue
        competitors.append({
            "brand":                str(row[0]).strip(),
            "positioning":          str(row[1]).strip() if row[1] else None,
            "competes_for":         str(row[2]).strip() if row[2] else None,
            "competitive_strength": str(row[3]).strip() if row[3] else None,
            "price_tier":           str(row[4]).strip() if row[4] else None,
        })
    return competitors


def _load_assumptions(wb) -> dict:
    """Assumptions & Model Settings: hardcode the 11 known rows by number for reliability.
    Each entry: {"value": parsed_value, "row": row_number, "label": str}"""
    ws = wb['Assumptions & Model Settings']

    row_defs = {
        5:  ("baseline_conversion_lift_low",      parse_pct),
        6:  ("baseline_conversion_lift_expected", parse_pct),
        7:  ("baseline_conversion_lift_high",     parse_pct),
        8:  ("organic_baseline_growth_rate",      parse_pct),
        9:  ("risk_adjustment",                   parse_pct),
        10: ("attribution_window_days",           lambda v: int(float(str(v))) if v is not None else None),
        17: ("total_campaign_budget",             parse_currency),
    }

    assumptions = {}
    for row_num, (key, parser) in row_defs.items():
        raw_val   = ws.cell(row=row_num, column=2).value
        raw_label = ws.cell(row=row_num, column=1).value
        assumptions[key] = {
            "value": parser(raw_val),
            "row":   row_num,
            "label": str(raw_label).strip() if raw_label else key,
        }

    # Scan-based: pick up cohort reach assumptions and budget split rows
    # by matching label substrings, so they work for any brand's cohort names.
    # Cohort reach rows are detected by 'reachable share' substring, keyed by cohort name.
    for row_num in range(1, ws.max_row + 1):
        label_cell = ws.cell(row=row_num, column=1).value
        if label_cell is None:
            continue
        label_str   = str(label_cell).strip()
        label_lower = label_str.lower()

        # Cohort reach: match 'reachable share' in label
        if 'reachable share' in label_lower:
            # Extract cohort name (everything before 'reachable share')
            # Patterns: "Classic Repeat: reachable share (%)" or "Classic Repeat — reachable share (%)"
            cohort_part = re.split(r'[:—\-]\s*reachable', label_str, flags=re.IGNORECASE)[0].strip()
            # Build key like "reach_classic_repeat"
            cohort_key  = re.sub(r'[^a-z0-9]+', '_', cohort_part.lower()).strip('_')
            key         = f"reach_{cohort_key}"
            raw_val     = ws.cell(row=row_num, column=2).value
            assumptions[key] = {
                "value": parse_pct(raw_val),
                "row":   row_num,
                "label": label_str,
                "cohort_name": cohort_part,
            }

        # Budget split rows
        if 'creative production share' in label_lower and 'creative_production_share' not in assumptions:
            raw_val = ws.cell(row=row_num, column=2).value
            assumptions["creative_production_share"] = {
                "value": parse_pct(raw_val),
                "row":   row_num,
                "label": label_str,
            }
        if 'media budget share' in label_lower and 'media_budget_share' not in assumptions:
            raw_val = ws.cell(row=row_num, column=2).value
            assumptions["media_budget_share"] = {
                "value": parse_pct(raw_val),
                "row":   row_num,
                "label": label_str,
            }

    return assumptions
